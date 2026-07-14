"""Geração de relatórios em Excel e PDF.

Ambos usam os mesmos dados do resumo do mês. Os arquivos são salvos em
data/relatorios/ e o caminho é devolvido para o app abrir.
"""

import os
from datetime import date, datetime

from .db import BASE_DIR
from . import repository as repo
from . import cartoes

REL_DIR = os.path.join(BASE_DIR, "data", "relatorios")


def _reais(cents):
    return round(cents / 100.0, 2)


def _fmt(cents):
    return f"R$ {cents/100:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _nome_mes(mes):
    meses = ["", "janeiro", "fevereiro", "março", "abril", "maio", "junho",
             "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    ano, m = mes.split("-")
    return f"{meses[int(m)]} de {ano}"


def _garantir_dir():
    os.makedirs(REL_DIR, exist_ok=True)


def _fmt_data_br(iso):
    a, m, d = iso.split("-")
    return f"{d}/{m}/{a}"


def _resolver_periodo(mes=None, data_inicio=None, data_fim=None):
    """Decide entre relatório mensal e por período livre.

    Retorna (resumo, transacoes, rotulo, sufixo_arquivo).
    """
    if data_inicio and data_fim:
        resumo = repo.resumo_periodo(data_inicio, data_fim)
        transacoes = repo.listar_transacoes(data_inicio=data_inicio, data_fim=data_fim)
        rotulo = f"{_fmt_data_br(data_inicio)} a {_fmt_data_br(data_fim)}"
        sufixo = f"{data_inicio}_a_{data_fim}"
    else:
        mes = mes or date.today().strftime("%Y-%m")
        resumo = repo.resumo_mes(mes)
        transacoes = repo.listar_transacoes(mes=mes)
        rotulo = _nome_mes(mes)
        sufixo = mes
    return resumo, transacoes, rotulo, sufixo


def _meses_no_intervalo(mes_ini, mes_fim):
    meses, atual = [], mes_ini
    while atual <= mes_fim:
        meses.append(atual)
        ano, m = map(int, atual.split("-"))
        m += 1
        ano += (m - 1) // 12
        m = (m - 1) % 12 + 1
        atual = f"{ano:04d}-{m:02d}"
    return meses


def _faturas_cartoes(mes=None, data_inicio=None, data_fim=None):
    """Faturas de cada cartão cujos meses caem no período do relatório.

    Retorna lista de dicts com nome do cartão, mês, total/pago/aberto e a lista
    de itens (compras) da fatura. Só inclui faturas que têm algum valor.
    """
    if data_inicio and data_fim:
        meses = _meses_no_intervalo(data_inicio[:7], data_fim[:7])
    else:
        meses = [mes or date.today().strftime("%Y-%m")]

    resultado = []
    for c in cartoes.listar_cartoes():
        for m in meses:
            f = cartoes.fatura(c["id"], m)
            if f["total_cents"] <= 0:
                continue
            resultado.append({
                "cartao": c["nome"], "mes": m,
                "total_cents": f["total_cents"], "pago_cents": f["pago_cents"],
                "aberto_cents": f["aberto_cents"], "paga": f["paga"],
                "itens": f["parcelas"],
            })
    return resultado


# ----------------------------------------------------------------------------
# EXCEL
# ----------------------------------------------------------------------------
def exportar_excel(mes=None, data_inicio=None, data_fim=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    _garantir_dir()
    resumo, transacoes, rotulo, sufixo = _resolver_periodo(mes, data_inicio, data_fim)

    verde = "1B7A5A"
    cabec_fill = PatternFill("solid", fgColor=verde)
    cabec_font = Font(bold=True, color="FFFFFF")
    titulo_font = Font(bold=True, size=16, color=verde)
    borda = Border(bottom=Side(style="thin", color="DDDDDD"))

    wb = Workbook()

    # --- Aba Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    ws["A1"] = "MeuCaixa — Relatório"
    ws["A1"].font = titulo_font
    ws["A2"] = rotulo
    ws["A2"].font = Font(size=12, color="6B7772")

    linhas_resumo = [
        ("Entradas", _reais(resumo["entradas_cents"])),
        ("Saídas", _reais(resumo["saidas_cents"])),
        ("Saldo do mês", _reais(resumo["saldo_mes_cents"])),
        ("Dinheiro em caixa", _reais(resumo["caixa_cents"])),
    ]
    r = 4
    for rot, val in linhas_resumo:
        ws[f"A{r}"] = rot
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = val
        ws[f"B{r}"].number_format = 'R$ #,##0.00'
        r += 1

    # tabela por categoria
    r += 1
    ws[f"A{r}"] = "Gastos por categoria"
    ws[f"A{r}"].font = Font(bold=True, size=12, color=verde)
    r += 1
    ws[f"A{r}"] = "Categoria"
    ws[f"B{r}"] = "Valor"
    for col in ("A", "B"):
        ws[f"{col}{r}"].fill = cabec_fill
        ws[f"{col}{r}"].font = cabec_font
    inicio_dados = r + 1
    for c in resumo["por_categoria"]:
        r += 1
        ws[f"A{r}"] = c["nome"]
        ws[f"B{r}"] = _reais(c["total"])
        ws[f"B{r}"].number_format = 'R$ #,##0.00'
    fim_dados = r

    # gráfico de barras (se houver dados)
    if fim_dados >= inicio_dados:
        chart = BarChart()
        chart.title = "Gastos por categoria"
        chart.type = "bar"
        chart.legend = None
        dados = Reference(ws, min_col=2, min_row=inicio_dados - 1, max_row=fim_dados)
        cats = Reference(ws, min_col=1, min_row=inicio_dados, max_row=fim_dados)
        chart.add_data(dados, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 16
        ws.add_chart(chart, f"D4")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16

    # --- Aba Transações ---
    ws2 = wb.create_sheet("Transações")
    cabecalhos = ["Data", "Descrição", "Categoria", "Tipo", "Valor (R$)", "Observação"]
    for i, h in enumerate(cabecalhos, start=1):
        cel = ws2.cell(row=1, column=i, value=h)
        cel.fill = cabec_fill
        cel.font = cabec_font
    for j, t in enumerate(transacoes, start=2):
        ws2.cell(row=j, column=1, value=t["data"])
        ws2.cell(row=j, column=2, value=t["descricao"])
        ws2.cell(row=j, column=3, value=t.get("categoria_nome") or "—")
        ws2.cell(row=j, column=4, value="Entrada" if t["tipo"] == "entrada" else "Saída")
        cel_val = ws2.cell(row=j, column=5, value=_reais(t["valor_cents"]))
        cel_val.number_format = 'R$ #,##0.00'
        ws2.cell(row=j, column=6, value=t.get("observacao") or "")
        for col in range(1, 7):
            ws2.cell(row=j, column=col).border = borda
    larguras = [12, 32, 22, 10, 14, 28]
    for i, w in enumerate(larguras, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # --- Aba Cartões (faturas do período) ---
    faturas = _faturas_cartoes(mes, data_inicio, data_fim)
    if faturas:
        ws3 = wb.create_sheet("Cartões")
        cab = ["Cartão", "Fatura (mês)", "Descrição", "Categoria", "Valor (R$)"]
        for i, h in enumerate(cab, start=1):
            cel = ws3.cell(row=1, column=i, value=h)
            cel.fill = cabec_fill
            cel.font = cabec_font
        linha = 2
        for f in faturas:
            for item in f["itens"]:
                ws3.cell(row=linha, column=1, value=f["cartao"])
                ws3.cell(row=linha, column=2, value=f["mes"])
                parc = f" ({item['numero']}/{item['parcelas_total']})" if item["parcelas_total"] > 1 else ""
                ws3.cell(row=linha, column=3, value=item["descricao"] + parc)
                ws3.cell(row=linha, column=4, value=item.get("categoria_nome") or "—")
                cel_val = ws3.cell(row=linha, column=5, value=_reais(item["valor_cents"]))
                cel_val.number_format = 'R$ #,##0.00'
                linha += 1
            # linha de subtotal da fatura
            ws3.cell(row=linha, column=3,
                     value=f"Total {f['cartao']} · {f['mes']} — pago {_fmt(f['pago_cents'])}, em aberto").font = Font(bold=True)
            cel_tot = ws3.cell(row=linha, column=5, value=_reais(f["aberto_cents"]))
            cel_tot.number_format = 'R$ #,##0.00'
            cel_tot.font = Font(bold=True)
            linha += 2
        larg3 = [22, 14, 34, 22, 14]
        for i, w in enumerate(larg3, start=1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    caminho = os.path.join(REL_DIR, f"MeuCaixa_{sufixo}.xlsx")
    wb.save(caminho)
    return caminho


# ----------------------------------------------------------------------------
# PDF
# ----------------------------------------------------------------------------
def exportar_pdf(mes=None, data_inicio=None, data_fim=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    )

    _garantir_dir()
    resumo, transacoes, rotulo, sufixo = _resolver_periodo(mes, data_inicio, data_fim)

    verde = colors.HexColor("#1B7A5A")
    cinza = colors.HexColor("#6B7772")
    claro = colors.HexColor("#EEF2EF")

    caminho = os.path.join(REL_DIR, f"MeuCaixa_{sufixo}.pdf")
    doc = SimpleDocTemplate(caminho, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], textColor=verde, fontSize=20)
    sub = ParagraphStyle("sub", parent=styles["Normal"], textColor=cinza, fontSize=11)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], textColor=verde, fontSize=13)
    normal = styles["Normal"]

    el = []
    el.append(Paragraph("MeuCaixa — Relatório", h1))
    el.append(Paragraph(rotulo, sub))
    el.append(Spacer(1, 0.6*cm))

    # cartões de resumo
    resumo_data = [[
        Paragraph("<b>Entradas</b><br/>" + _fmt(resumo["entradas_cents"]), normal),
        Paragraph("<b>Saídas</b><br/>" + _fmt(resumo["saidas_cents"]), normal),
        Paragraph("<b>Saldo do mês</b><br/>" + _fmt(resumo["saldo_mes_cents"]), normal),
        Paragraph("<b>Em caixa</b><br/>" + _fmt(resumo["caixa_cents"]), normal),
    ]]
    t_resumo = Table(resumo_data, colWidths=[4.1*cm]*4)
    t_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), claro),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.white),
        ("INNERGRID", (0, 0), (-1, -1), 4, colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    el.append(t_resumo)
    el.append(Spacer(1, 0.7*cm))

    # gastos por categoria
    el.append(Paragraph("Gastos por categoria", h2))
    el.append(Spacer(1, 0.2*cm))
    if resumo["por_categoria"]:
        dados = [["Categoria", "Valor", "% das saídas"]]
        total_saidas = max(resumo["saidas_cents"], 1)
        for c in resumo["por_categoria"]:
            pct = c["total"] / total_saidas * 100
            dados.append([c["nome"], _fmt(c["total"]), f"{pct:.0f}%"])
        tab = Table(dados, colWidths=[8*cm, 4*cm, 4*cm])
        tab.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), verde),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, claro]),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        el.append(tab)
    else:
        el.append(Paragraph("Sem gastos registrados neste mês.", normal))
    el.append(Spacer(1, 0.7*cm))

    # transações
    el.append(Paragraph("Movimentações", h2))
    el.append(Spacer(1, 0.2*cm))
    if transacoes:
        dados = [["Data", "Descrição", "Categoria", "Valor"]]
        for t in transacoes[:60]:
            sinal = "+" if t["tipo"] == "entrada" else "−"
            dados.append([
                t["data"][8:10] + "/" + t["data"][5:7],
                t["descricao"][:34],
                (t.get("categoria_nome") or "—")[:20],
                f"{sinal} {_fmt(t['valor_cents'])}",
            ])
        tab = Table(dados, colWidths=[2*cm, 6.5*cm, 4.5*cm, 3*cm])
        tab.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), verde),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, claro]),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        el.append(tab)
    else:
        el.append(Paragraph("Nenhuma movimentação neste mês.", normal))

    # cartões de crédito (faturas do período)
    faturas = _faturas_cartoes(mes, data_inicio, data_fim)
    if faturas:
        el.append(Spacer(1, 0.7*cm))
        el.append(Paragraph("Cartões de crédito", h2))
        el.append(Spacer(1, 0.2*cm))
        dados = [["Cartão", "Fatura", "Compra", "Valor"]]
        for f in faturas:
            for item in f["itens"]:
                parc = f" ({item['numero']}/{item['parcelas_total']})" if item["parcelas_total"] > 1 else ""
                dados.append([f["cartao"][:16], f["mes"], (item["descricao"] + parc)[:30],
                              _fmt(item["valor_cents"])])
            situacao = "quitada" if f["paga"] else f"em aberto {_fmt(f['aberto_cents'])}"
            dados.append(["", "", f"Total {f['mes']} ({situacao})", _fmt(f["total_cents"])])
        tab = Table(dados, colWidths=[3.5*cm, 2.2*cm, 7.3*cm, 3*cm])
        estilo = [
            ("BACKGROUND", (0, 0), (-1, 0), verde),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (3, 0), (3, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        # negrito nas linhas de subtotal (as que começam com "Total")
        for i, linha in enumerate(dados):
            if linha[2].startswith("Total"):
                estilo.append(("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"))
                estilo.append(("LINEABOVE", (0, i), (-1, i), 0.5, cinza))
        tab.setStyle(TableStyle(estilo))
        el.append(tab)

    el.append(Spacer(1, 1*cm))
    el.append(Paragraph(
        f"Gerado pelo MeuCaixa em {datetime.now().strftime('%d/%m/%Y %H:%M')}.", sub))

    doc.build(el)
    return caminho
